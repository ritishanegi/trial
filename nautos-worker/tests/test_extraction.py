"""
Tests for ExtractionService.

Test conventions (matching the existing test_privacy.py pattern):
  - Plain pytest functions in nautos-worker/tests/
  - unittest.mock (no pytest-mock in pyproject.toml dev deps)
  - No conftest.py fixtures needed — services are stateless and easily constructed
  - Integration tests are gated behind @pytest.mark.integration and skipped by default

Assumption corrections vs. the original spec:
  - Document ID key:    "document_id" (confirmed from embeddings table schema)
  - Chunk ordering:     ORDER BY page_number ASC, chunk_index ASC (confirmed in VectorDBService)
  - Return type:        ExtractionService.extract_to_excel returns (bytes, filename), NOT a file path
  - Exception type:     ValueError (not a custom ExtractionError) — consistent with the codebase's
                        _PERMANENT_EXC_TYPES list in tasks/ingestion.py
  - Test directory:     tests/ at repo root (not app/tests/) — matches existing test_privacy.py
  - LLM call pattern:   llm.get_answer(question=..., context="") — confirmed from extraction.py
  - VectorDB method:    vectordb.get_all_chunks_for_document(document_id=..., tenant_id=...)
"""

import io
import json
import os
from unittest.mock import MagicMock, patch, call

import pytest
from openpyxl import load_workbook

from app.services.extraction import ExtractionService


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_chunk(text: str, page: int = 1, chunk_index: int = 0, title: str = "Test Doc") -> dict:
    """Build a chunk dict matching VectorDBService.get_all_chunks_for_document output."""
    return {
        "document_id": "doc-123",
        "text": text,
        "page_number": page,
        "chunk_index": chunk_index,
        "score": 1.0,
        "scope": "vessel",
        "title": title,
    }


def _make_service(mock_vectordb=None, mock_llm=None) -> ExtractionService:
    """
    Build an ExtractionService with both external dependencies replaced by mocks.
    Patches at construction time so no real DB/LLM connections are attempted.
    """
    svc = ExtractionService.__new__(ExtractionService)
    svc.vectordb = mock_vectordb or MagicMock()
    svc.llm = mock_llm or MagicMock()
    return svc


def _load_xlsx(xlsx_bytes: bytes):
    """Load an openpyxl workbook from raw bytes."""
    return load_workbook(io.BytesIO(xlsx_bytes))


# ── Task 1: fetch_document_chunks (via get_all_chunks_for_document) ────────────

class TestFetchDocumentChunks:
    """
    The spec's fetch_document_chunks maps directly to:
    VectorDBService.get_all_chunks_for_document(document_id, tenant_id)
    The result is consumed inside extract_to_excel.
    We test the vectordb call and ordering assumptions here.
    """

    def test_fetch_document_chunks_success(self):
        """Mock vectordb to return a fixed list; assert the service uses them in order."""
        chunks = [
            _make_chunk("First chunk", page=1, chunk_index=0),
            _make_chunk("Second chunk", page=1, chunk_index=1),
            _make_chunk("Third chunk", page=2, chunk_index=0),
        ]
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = chunks

        mock_llm = MagicMock()
        mock_llm.get_answer.return_value = '[{"part_no": "123", "qty": "1"}]'

        svc = _make_service(mock_vdb, mock_llm)
        xlsx_bytes, _ = svc.extract_to_excel("doc-123", "tenant-1", "parts list")

        # vectordb was called with the right IDs
        mock_vdb.get_all_chunks_for_document.assert_called_once_with(
            document_id="doc-123", tenant_id="tenant-1"
        )
        # Result is a valid xlsx
        wb = _load_xlsx(xlsx_bytes)
        assert wb.active is not None

    def test_fetch_document_chunks_empty(self):
        """Empty chunk list should raise ValueError, not crash with an unhandled exception."""
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = []

        svc = _make_service(mock_vdb)
        with pytest.raises(ValueError, match="no indexed content"):
            svc.extract_to_excel("doc-empty", "tenant-1", "parts list")

    def test_chunks_ordered_by_page_then_index(self):
        """
        Ordering is enforced in the SQL (ORDER BY page_number ASC, chunk_index ASC).
        Here we verify the service stitches context in the order the vectordb returns them,
        i.e. it does NOT re-sort — trusting the DB's guarantee.
        """
        chunks = [
            _make_chunk("Page 1 chunk A", page=1, chunk_index=0),
            _make_chunk("Page 1 chunk B", page=1, chunk_index=1),
            _make_chunk("Page 2 chunk A", page=2, chunk_index=0),
        ]
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = chunks

        captured_context = {}

        def capture_get_answer(question, context, **kwargs):
            captured_context["prompt"] = question
            return '[{"col": "val"}]'

        mock_llm = MagicMock()
        mock_llm.get_answer.side_effect = capture_get_answer

        svc = _make_service(mock_vdb, mock_llm)
        svc.extract_to_excel("doc-123", "tenant-1", "test")

        # Context must appear in page order
        prompt = captured_context["prompt"]
        idx_p1a = prompt.index("Page 1 chunk A")
        idx_p1b = prompt.index("Page 1 chunk B")
        idx_p2a = prompt.index("Page 2 chunk A")
        assert idx_p1a < idx_p1b < idx_p2a


# ── Task 2: extract_structured_fields (via _call_llm_and_parse) ───────────────

class TestExtractStructuredFields:

    def test_valid_json_parsed_correctly(self):
        """LLM returns clean JSON → parsed into list[dict] and written to xlsx."""
        raw_json = json.dumps([
            {"vessel_name": "MV Pacific Star", "imo_number": "9876543"},
            {"vessel_name": "MV Atlantic Hope", "imo_number": "1234567"},
        ])
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = [_make_chunk("some text")]

        mock_llm = MagicMock()
        mock_llm.get_answer.return_value = raw_json

        svc = _make_service(mock_vdb, mock_llm)
        xlsx_bytes, filename = svc.extract_to_excel("doc-123", "tenant-1", "vessel registry")

        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        # Headers in row 1
        headers = [ws.cell(1, c).value for c in range(1, 3)]
        assert "vessel_name" in headers
        assert "imo_number" in headers
        # Two data rows
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_valid_json_in_code_fence_parsed_correctly(self):
        """LLM wraps JSON in ```json fences — service must strip and parse successfully."""
        inner = '[{"part_no": "A1", "qty": "5"}]'
        fenced = f"```json\n{inner}\n```"

        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = [_make_chunk("text")]

        mock_llm = MagicMock()
        mock_llm.get_answer.return_value = fenced

        svc = _make_service(mock_vdb, mock_llm)
        xlsx_bytes, _ = svc.extract_to_excel("doc-123", "tenant-1", "parts")

        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        assert ws.cell(1, 1).value == "part_no"
        assert ws.cell(2, 1).value == "A1"

    def test_malformed_json_triggers_retry(self):
        """
        First LLM call returns garbage; second call returns valid JSON.
        Assert get_answer is called exactly twice and the second result is used.
        """
        valid_json = '[{"col": "val"}]'
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = [_make_chunk("text")]

        mock_llm = MagicMock()
        mock_llm.get_answer.side_effect = [
            "This is definitely not JSON.",  # first call — invalid
            valid_json,                       # second call — valid
        ]

        svc = _make_service(mock_vdb, mock_llm)
        xlsx_bytes, _ = svc.extract_to_excel("doc-123", "tenant-1", "test extract")

        assert mock_llm.get_answer.call_count == 2

        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        assert ws.cell(1, 1).value == "col"
        assert ws.cell(2, 1).value == "val"

    def test_persistent_json_failure_raises_value_error(self):
        """
        Both LLM calls return non-JSON.
        Assert ValueError is raised after retry exhaustion (not a silent empty result).
        """
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = [_make_chunk("text")]

        mock_llm = MagicMock()
        mock_llm.get_answer.side_effect = [
            "Not JSON at all.",
            "Still not JSON.",
        ]

        svc = _make_service(mock_vdb, mock_llm)

        with pytest.raises(ValueError, match="did not return valid JSON"):
            svc.extract_to_excel("doc-123", "tenant-1", "parts")

        # Confirm both attempts were made
        assert mock_llm.get_answer.call_count == 2

    def test_retry_prompt_contains_correction_instruction(self):
        """The retry prompt must include a corrective message nudging the model."""
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = [_make_chunk("text")]

        call_prompts = []

        def capture(question, context, **kwargs):
            call_prompts.append(question)
            # First call invalid, second valid
            if len(call_prompts) == 1:
                return "not json"
            return '[{"x": "1"}]'

        mock_llm = MagicMock()
        mock_llm.get_answer.side_effect = capture

        svc = _make_service(mock_vdb, mock_llm)
        svc.extract_to_excel("doc-123", "tenant-1", "test")

        assert len(call_prompts) == 2
        # The corrective instruction must appear in the retry prompt but NOT in the first
        assert "not valid JSON" in call_prompts[1]
        assert "not valid JSON" not in call_prompts[0]
        # The retry prompt appends the correction to the original — original must still be there
        assert "Output the JSON array now:" in call_prompts[1]
        # The correction suffix is appended at the end
        assert call_prompts[1].endswith(
            "Do not include any other text, code fences, or explanation."
        )


# ── Task 3: build_excel_workbook (via _build_xlsx) ───────────────────────────

class TestBuildExcelWorkbook:

    def _extract_with_rows(self, rows: list[dict]) -> bytes:
        """Helper: bypass LLM/vectordb and call _build_xlsx directly."""
        svc = ExtractionService.__new__(ExtractionService)
        return svc._build_xlsx(rows, sheet_name="TestSheet")

    def test_single_document_correct_headers_and_values(self):
        """Single document rows → correct column headers in row 1, data in row 2."""
        rows = [
            {"vessel_name": "MV Pacific Star", "imo_number": "9876543", "flag": "Panama"},
        ]
        xlsx_bytes = self._extract_with_rows(rows)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active

        headers = {ws.cell(1, c).value for c in range(1, 4)}
        assert {"vessel_name", "imo_number", "flag"} == headers

        # Find column index for vessel_name
        col_map = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, col_map["vessel_name"]).value == "MV Pacific Star"
        assert ws.cell(2, col_map["imo_number"]).value == "9876543"

    def test_header_row_is_bold(self):
        """Header row cells must have bold font — visual requirement from spec."""
        rows = [{"col_a": "val1"}]
        xlsx_bytes = self._extract_with_rows(rows)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is True

    def test_freeze_panes_on_a2(self):
        """Header row must be frozen at A2."""
        rows = [{"col_a": "val1"}]
        xlsx_bytes = self._extract_with_rows(rows)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_missing_fields_across_documents_produce_blank_cells(self):
        """
        Two documents with non-overlapping fields:
        - Doc A has field_a only
        - Doc B has field_b only
        Union of columns should be created; missing cells must be None (blank), not errored.
        """
        rows = [
            {"field_a": "value_a"},         # Doc A
            {"field_b": "value_b"},         # Doc B — missing field_a
        ]
        xlsx_bytes = self._extract_with_rows(rows)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active

        # Both columns must exist in header
        headers = {ws.cell(1, c).value for c in range(1, ws.max_column + 1)}
        assert "field_a" in headers
        assert "field_b" in headers

        col_map = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

        # Row 2 (Doc A): field_a=value_a, field_b=blank
        assert ws.cell(2, col_map["field_a"]).value == "value_a"
        assert ws.cell(2, col_map["field_b"]).value is None

        # Row 3 (Doc B): field_a=blank, field_b=value_b
        assert ws.cell(3, col_map["field_a"]).value is None
        assert ws.cell(3, col_map["field_b"]).value == "value_b"

    def test_column_order_follows_first_row_discovery(self):
        """
        Columns appear in the order keys are first seen across all rows.
        First row establishes col_a first; col_b only appears in second row.
        """
        rows = [
            {"col_a": "1", "col_b": "2"},
            {"col_c": "3"},
        ]
        xlsx_bytes = self._extract_with_rows(rows)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active

        col_order = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        # col_a and col_b appear before col_c (first-seen ordering)
        assert col_order.index("col_a") < col_order.index("col_c")
        assert col_order.index("col_b") < col_order.index("col_c")

    def test_dict_values_serialised_to_json_string(self):
        """
        openpyxl cannot write dicts/lists directly.
        The service must json.dumps() them — assert the cell contains a string, not an error.
        """
        rows = [{"nested": {"a": 1, "b": [2, 3]}}]
        xlsx_bytes = self._extract_with_rows(rows)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        cell_val = ws.cell(2, 1).value
        assert isinstance(cell_val, str)
        assert json.loads(cell_val) == {"a": 1, "b": [2, 3]}

    def test_column_width_capped_at_52(self):
        """
        Column widths are auto-sized but capped at 50 chars + 2 padding = 52.
        A very long value must not produce a column wider than 52.
        """
        rows = [{"col": "x" * 200}]
        xlsx_bytes = self._extract_with_rows(rows)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        col_letter = ws.cell(1, 1).column_letter
        width = ws.column_dimensions[col_letter].width
        assert width <= 52


# ── Task 4: end-to-end orchestration (extract_to_excel) ──────────────────────

class TestRunEndToEnd:

    def test_end_to_end_with_mocks_produces_valid_xlsx(self):
        """
        Full pipeline with mocked vectordb + LLM.
        Assert:
          - A non-empty bytes object is returned
          - The bytes are a valid .xlsx
          - The xlsx contains the expected headers and data
        """
        chunks = [
            _make_chunk("Part A  PN:001  Qty:5  Location:Engine Room", page=1),
            _make_chunk("Part B  PN:002  Qty:2  Location:Bridge", page=2),
        ]
        llm_response = json.dumps([
            {"part_no": "001", "description": "Part A", "qty": "5", "location": "Engine Room"},
            {"part_no": "002", "description": "Part B", "qty": "2", "location": "Bridge"},
        ])

        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = chunks

        mock_llm = MagicMock()
        mock_llm.get_answer.return_value = llm_response

        svc = _make_service(mock_vdb, mock_llm)
        xlsx_bytes, filename = svc.extract_to_excel(
            document_id="doc-e2e",
            tenant_id="tenant-1",
            description="parts list with part number, description, qty, location",
        )

        # Return types
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0
        assert filename.endswith(".xlsx")

        # Content
        wb = _load_xlsx(xlsx_bytes)
        ws = wb.active
        headers = {ws.cell(1, c).value for c in range(1, ws.max_column + 1)}
        assert {"part_no", "description", "qty", "location"} == headers
        assert ws.max_row == 3  # header + 2 data rows

    def test_filename_is_derived_from_doc_title_and_description(self):
        """Returned filename should contain a sanitised form of the document title."""
        chunks = [_make_chunk("text", title="Engine Manual 2024")]
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = chunks

        mock_llm = MagicMock()
        mock_llm.get_answer.return_value = '[{"col": "val"}]'

        svc = _make_service(mock_vdb, mock_llm)
        _, filename = svc.extract_to_excel("doc-123", "tenant-1", "spare parts")

        assert "Engine_Manual_2024" in filename
        assert filename.endswith(".xlsx")

    def test_empty_document_raises_before_calling_llm(self):
        """If vectordb returns no chunks, the LLM must not be called at all."""
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = []

        mock_llm = MagicMock()

        svc = _make_service(mock_vdb, mock_llm)

        with pytest.raises(ValueError, match="no indexed content"):
            svc.extract_to_excel("doc-empty", "tenant-1", "parts")

        mock_llm.get_answer.assert_not_called()

    def test_llm_returns_empty_array_raises(self):
        """An empty JSON array from the LLM should raise ValueError (no rows to export)."""
        mock_vdb = MagicMock()
        mock_vdb.get_all_chunks_for_document.return_value = [_make_chunk("text")]

        mock_llm = MagicMock()
        mock_llm.get_answer.return_value = "[]"

        svc = _make_service(mock_vdb, mock_llm)

        with pytest.raises(ValueError, match="did not find any matching rows"):
            svc.extract_to_excel("doc-123", "tenant-1", "parts")


# ── Safe-name helpers ─────────────────────────────────────────────────────────

class TestSafeHelpers:

    def test_safe_sheet_name_truncates_at_31(self):
        long_name = "A" * 40
        result = ExtractionService._safe_sheet_name(long_name)
        assert len(result) <= 31

    def test_safe_sheet_name_strips_illegal_chars(self):
        result = ExtractionService._safe_sheet_name("Parts / List [2024]")
        for ch in r"/\*?:[]":
            assert ch not in result

    def test_safe_sheet_name_fallback_on_empty(self):
        result = ExtractionService._safe_sheet_name("///")
        assert result == "Sheet1"

    def test_safe_filename_replaces_special_chars(self):
        result = ExtractionService._safe_filename("Engine Manual 2024!")
        assert " " not in result
        assert "!" not in result


# ── Integration test (skipped unless NAUTOS_INTEGRATION_TESTS=1) ──────────────

@pytest.mark.skipif(
    os.environ.get("NAUTOS_INTEGRATION_TESTS") != "1",
    reason=(
        "Integration test: requires a running PostgreSQL+pgvector DB with a real ingested "
        "document and a configured LLM API key. "
        "Enable with: NAUTOS_INTEGRATION_TESTS=1 pytest tests/test_extraction.py -m integration"
    ),
)
@pytest.mark.integration
def test_integration_real_document():
    """
    Full end-to-end integration test against a real ingested document.

    Prerequisites:
      1. Start the dev stack: make up
      2. Ingest a sample PDF (see MANUAL_TESTING.md for step-by-step)
      3. Set env vars:
           NAUTOS_INTEGRATION_TESTS=1
           INTEGRATION_TEST_DOCUMENT_ID=<id from the ingested document>
           INTEGRATION_TEST_TENANT_ID=<matching tenant_id>

    Asserts:
      - Returned bytes are a valid .xlsx
      - The workbook has at least 1 data row
      - At least N expected columns exist (adjust based on your sample PDF)
    """
    doc_id = os.environ.get("INTEGRATION_TEST_DOCUMENT_ID")
    tenant_id = os.environ.get("INTEGRATION_TEST_TENANT_ID", "default")
    min_expected_columns = int(os.environ.get("INTEGRATION_TEST_MIN_COLUMNS", "3"))

    if not doc_id:
        pytest.skip("INTEGRATION_TEST_DOCUMENT_ID not set")

    svc = ExtractionService()
    xlsx_bytes, filename = svc.extract_to_excel(
        document_id=doc_id,
        tenant_id=tenant_id,
        description="parts list with part number, designation, and quantity",
    )

    assert isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0
    assert filename.endswith(".xlsx")

    wb = _load_xlsx(xlsx_bytes)
    ws = wb.active
    assert ws.max_row >= 2, "Expected at least 1 data row in the output"
    assert ws.max_column >= min_expected_columns, (
        f"Expected at least {min_expected_columns} columns, got {ws.max_column}"
    )
